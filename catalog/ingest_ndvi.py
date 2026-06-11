"""Ingest the CropWise whole-farm daily-NDVI export (vegetation_history_daily.xlsx)
into vegetation_weekly for all fields.

The file is WIDE: column 0 = date, column 1 = day-number, columns 2+ = one column
per field (header = the field's "Имя"). We map each column to a field_id (via the
field name), downsample daily NDVI to the ISO-week mean, and upsert as
source='cropwise_bulk'. Then we drop the older single-field NDVI rows so the
baseline + card read one consistent farm-wide source.

Run: docker compose -f docker-compose.prod.yml run --rm -T -v /tmp/veg.xlsx:/data.xlsx \
       bot python -m catalog.ingest_ndvi /data.xlsx
"""
import argparse
import asyncio
import re
import sys
from datetime import date, datetime, timedelta

import openpyxl
from sqlalchemy import text

from bot.db import engine


def _ima(name):
    """Recover the CropWise 'Имя' from our field name."""
    if " · " in name:                       # "Поле 25 · Красное" -> "25"
        return name.split(" · ")[0].replace("Поле ", "").strip()
    m = re.match(r"Поле\s+(\d+)", name or "")  # pilot "Поле 76/108" -> "76"
    return m.group(1) if m else (name or "").strip()


def _date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip().split(" ")[0]
    for f in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, f).date()
        except ValueError:
            pass
    return None


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", "."))
    except ValueError:
        return None


async def run(path):
    async with engine.begin() as conn:
        fields = (await conn.execute(text("SELECT id, name FROM fields"))).all()
        ima2id = {}
        for fid, nm in fields:
            ima2id.setdefault(_ima(nm), fid)

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = list(next(it))
        col2fid, unmapped = {}, []
        for ci in range(2, len(header)):
            h = header[ci]
            if h is None:
                continue
            fid = ima2id.get(str(h).strip())
            if fid:
                col2fid[ci] = fid
            else:
                unmapped.append(str(h).strip())

        agg = {}  # (fid, isoyear, isoweek) -> [sum, count, monday]
        for row in it:
            d = _date(row[0])
            if not d:
                continue
            iy, iw, iwd = d.isocalendar()
            monday = d - timedelta(days=iwd - 1)
            for ci, fid in col2fid.items():
                if ci < len(row):
                    v = _num(row[ci])
                    if v is not None:
                        k = (fid, iy, iw)
                        a = agg.get(k)
                        if a:
                            a[0] += v; a[1] += 1
                        else:
                            agg[k] = [v, 1, monday]
        wb.close()

        rows = [{"f": fid, "ws": mon, "wn": iw, "nd": round(s / c, 4)}
                for (fid, _iy, iw), (s, c, mon) in agg.items()]
        ins = text("INSERT INTO vegetation_weekly (field_id, week_start, week_no, ndvi, source) "
                   "VALUES (:f,:ws,:wn,:nd,'cropwise_bulk') "
                   "ON CONFLICT (field_id, week_start, source) "
                   "DO UPDATE SET ndvi=EXCLUDED.ndvi, week_no=EXCLUDED.week_no")
        for i in range(0, len(rows), 1000):
            await conn.execute(ins, rows[i:i + 1000])
        deleted = (await conn.execute(
            text("DELETE FROM vegetation_weekly WHERE source <> 'cropwise_bulk'"))).rowcount

    print(f"mapped {len(col2fid)} field-columns; {len(unmapped)} unmapped "
          f"({', '.join(unmapped[:12])}{'…' if len(unmapped) > 12 else ''})", file=sys.stderr)
    print(f"loaded {len(rows)} weekly NDVI rows; dropped {deleted} old single-field rows.",
          file=sys.stderr)
    return 0


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    args = ap.parse_args()
    return asyncio.run(run(args.xlsx))


if __name__ == "__main__":
    sys.exit(main())
