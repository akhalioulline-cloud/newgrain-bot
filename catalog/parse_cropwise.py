"""Parse a CropWise field 'multiprotocol' .xlsx export into a normalized CSV
that catalog/ingest_treatments.py can load into field_treatments.

CropWise exports one sheet per year ('Технические операции YYYY'), each split
into sections: Обработка почвы (tillage) / Сев (sowing) / Внесение удобрений
(fertilizer) / Внесение СЗР (plant protection). We read each section
header-driven (robust to column order) and emit one normalized row per operation.

Runs LOCALLY (needs openpyxl); produces a ';'-delimited CSV. Then load with
ingest_treatments.py on the server.

Single file:
    python catalog/parse_cropwise.py <multiprotocol.xlsx> <out.csv> --field "Поле 76/108"
Whole folder (one combined CSV for all fields at once) — the field is read per
file from its 'Паспорт поля' sheet (number + group + area), or pinned with
--field-map:
    python catalog/parse_cropwise.py <folder/> <out.csv> [--field-map map.csv]
Every row is tagged source='cropwise_multiprotocol'; ingest_treatments dedups on
the natural key, so re-runs and a later CropWise-API sync never double-count.
"""
import argparse
import csv
import io
import os
import re
import sys

import openpyxl

MONTHS = {"января": 1, "февраля": 2, "марта": 3, "апреля": 4, "мая": 5, "июня": 6,
          "июля": 7, "августа": 8, "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12}
SECTIONS = {  # section header label -> (category, material-column candidates)
    "Обработка почвы": ("tillage", []),
    "Сев": ("sowing", ["Семена"]),
    "Внесение удобрений": ("fertilizer", ["Удобрение"]),
    "Внесение СЗР": ("protection", ["СЗР"]),
}
OUT_COLS = ["поле", "дата", "культура", "операция", "категория", "препарат",
            "норма", "площадь_га", "примечание", "источник", "площадь_поля"]
DEFAULT_SOURCE = "cropwise_multiprotocol"


def _round(s):
    try:
        return "%g" % round(float(str(s).replace(",", ".")), 3)
    except (ValueError, TypeError):
        return str(s or "").strip()


def _ru_date(s):
    # "05 мая 2022, 13:00" -> "05.05.2022"
    m = re.match(r"\s*(\d{1,2})\s+([а-яё]+)\s+(\d{4})", str(s or ""), re.I)
    if not m:
        return ""
    d, mon, y = m.group(1), m.group(2).lower(), m.group(3)
    mn = MONTHS.get(mon)
    return f"{int(d):02d}.{mn:02d}.{y}" if mn else ""


def _crop(op_name, sheet_default):
    t = (op_name or "").lower()
    if "сои" in t or "сою" in t or " соя" in t or t.startswith("соя"):
        return "Соя"
    if "подсолнеч" in t:
        return "Подсолнечник"
    if "озим" in t and "пшениц" in t:
        return "Озимая пшеница"
    if "яров" in t and "пшениц" in t:
        return "Яровая пшеница"
    if "пшениц" in t:
        return "Пшеница"
    return sheet_default


def _sheet_default_crop(ops):
    blob = " ".join(ops).lower()
    for key, crop in [("сои", "Соя"), ("сою", "Соя"), ("подсолнеч", "Подсолнечник"),
                      ("озим", "Озимая пшеница"), ("яров", "Яровая пшеница"), ("пшениц", "Пшеница")]:
        if key in blob:
            return crop
    return ""


def parse(path, field_name, source=DEFAULT_SOURCE, field_area=None):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for sh in wb.sheetnames:
        if not sh.startswith("Технические операции"):
            continue
        rows = [list(r) for r in wb[sh].iter_rows(values_only=True)]
        # first pass: collect operation names to infer the sheet's dominant crop
        op_names = [str(r[2]) for r in rows if len(r) > 2 and r[2] and r[0] not in (None, "№")]
        sheet_crop = _sheet_default_crop(op_names)

        section = None
        hdr = {}
        for r in rows:
            c0 = (str(r[0]).strip() if r and r[0] is not None else "")
            if c0 in SECTIONS:
                section, hdr = c0, {}
                continue
            if c0 == "№":  # header row -> map column name -> index
                hdr = {str(c).strip(): i for i, c in enumerate(r) if c}
                continue
            if not section or not c0.isdigit() or not hdr:
                continue

            def cell(name):
                i = hdr.get(name)
                return r[i] if i is not None and i < len(r) and r[i] is not None else ""

            cat, mat_cols = SECTIONS[section]
            date = _ru_date(cell("Date"))
            op = str(cell("Название технической операции")).strip()
            product = ""
            for mc in mat_cols:
                if cell(mc):
                    product = str(cell(mc)).strip(); break
            rate = _round(cell("Норма (факт)"))
            unit = str(cell("Единицы")).strip()
            total = _round(cell("Всего (факт)"))
            area = str(cell("Площадь, га")).strip()
            depth = str(cell("Глубина")).strip()
            norma = f"{rate} {unit}".strip() if rate else ""
            note_bits = []
            if total:
                note_bits.append(f"всего {total}{(' ' + unit) if unit else ''}")
            if depth:
                note_bits.append(f"глубина {depth}")
            out.append({
                "поле": field_name,
                "дата": date,
                "культура": _crop(op, sheet_crop),
                "операция": op,
                "категория": cat,
                "препарат": product,
                "норма": norma,
                "площадь_га": area,
                "примечание": "; ".join(note_bits),
                "источник": source,
                "площадь_поля": field_area,
            })
    wb.close()
    return out


def _cell_str(v):
    """Stringify a passport value; render integer-valued floats without '.0'."""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _passport(path):
    """Read the field's number, group and official area from the 'Паспорт поля'
    sheet that CropWise puts first in every multiprotocol export. Returns the
    fields-table display name ('Поле <номер> · <группа>', matching
    ingest_fields) plus the area for number+area fallback matching."""
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return None, None
    if "Паспорт поля" not in wb.sheetnames:
        wb.close()
        return None, None
    num = group = area = None
    for row in wb["Паспорт поля"].iter_rows(values_only=True):
        cells = list(row)
        for i, c in enumerate(cells):
            label = str(c).strip() if c is not None else ""
            nxt = cells[i + 1] if i + 1 < len(cells) else None
            if nxt in (None, ""):
                continue
            if label == "Поле" and num is None:
                num = _cell_str(nxt)
            elif label == "Группа полей" and group is None:
                group = _cell_str(nxt)
            elif label.startswith("Площадь") and area is None:
                area = _cell_str(nxt)
        if num and group and area:
            break
    wb.close()
    if not num:
        return None, None
    name = f"Поле {num}" + (f" · {group}" if group else "")
    return name, area


def _field_token(s):
    """Last resort: pull a «76/108»-style field number out of a string."""
    m = re.search(r"(\d+)\s*[/\-]\s*(\d+)", str(s or ""))
    return f"Поле {m.group(1)}/{m.group(2)}" if m else None


def _load_field_map(path):
    """CSV of `filename,field_name` (or `cropwise_id,field_name`) — the
    deterministic override for batch runs when auto-detection isn't reliable."""
    out = {}
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "cp1251", "utf-8"):
        try:
            data = raw.decode(enc)
            break
        except UnicodeDecodeError:
            continue
    delim = ";" if data[:2000].count(";") > data[:2000].count(",") else ","
    for row in csv.reader(io.StringIO(data), delimiter=delim):
        if len(row) >= 2 and row[0].strip():
            out[row[0].strip()] = row[1].strip()
    return out


def _resolve_field(path, explicit, field_map):
    """Return (field_name, field_area). Priority: --field-map override →
    explicit --field → the 'Паспорт поля' sheet → a token in the filename."""
    base = os.path.basename(path)
    if field_map:
        if base in field_map:
            return field_map[base], None
        tok = re.search(r"\d+", base)  # match by a bare id in the filename
        if tok and tok.group(0) in field_map:
            return field_map[tok.group(0)], None
    if explicit:
        return explicit, None
    name, area = _passport(path)
    if name:
        return name, area
    return _field_token(base), None


def main():
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("xlsx", help="a multiprotocol .xlsx OR a folder of them")
    ap.add_argument("out_csv", help="combined output CSV")
    ap.add_argument("--field", default=None,
                    help="field name (single file). Omit in folder mode — "
                         "the field is auto-detected per file or via --field-map.")
    ap.add_argument("--field-map", default=None,
                    help="CSV filename,field_name to pin each file's field")
    ap.add_argument("--source", default=DEFAULT_SOURCE,
                    help=f"source tag stored on every row (default {DEFAULT_SOURCE})")
    args = ap.parse_args()

    field_map = _load_field_map(args.field_map) if args.field_map else None

    if os.path.isdir(args.xlsx):
        files = sorted(
            os.path.join(args.xlsx, f) for f in os.listdir(args.xlsx)
            if f.lower().endswith(".xlsx") and not f.startswith("~$")
        )
        if not files:
            print(f"no .xlsx files in {args.xlsx}", file=sys.stderr)
            return 1
    else:
        files = [args.xlsx]

    rows, unresolved = [], []
    for path in files:
        field, area = _resolve_field(path, args.field, field_map)
        if not field:
            unresolved.append(os.path.basename(path))
            print(f"  ⚠ could not resolve field for {os.path.basename(path)} — "
                  f"skipped (use --field-map)", file=sys.stderr)
            continue
        frows = [r for r in parse(path, field, args.source, area) if r["дата"]]
        rows.extend(frows)
        ha = f" ({area} га)" if area else ""
        print(f"  {os.path.basename(path)} -> {field}{ha}: {len(frows)} ops",
              file=sys.stderr)

    if not rows:
        print("no datable ops parsed.", file=sys.stderr)
        return 1

    with open(args.out_csv, "w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=OUT_COLS, delimiter=";")
        w.writeheader()
        w.writerows(rows)

    cats, fields = {}, set()
    for r in rows:
        cats[r["категория"]] = cats.get(r["категория"], 0) + 1
        fields.add(r["поле"])
    print(f"wrote {len(rows)} ops across {len(fields)} field(s) -> {args.out_csv}",
          file=sys.stderr)
    print(f"by category: {cats}", file=sys.stderr)
    if unresolved:
        print(f"{len(unresolved)} file(s) skipped (unresolved field): "
              f"{', '.join(unresolved)}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
