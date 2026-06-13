"""Export a field's operations as a CropWise-style 'multiprotocol' .xlsx.

The inverse of catalog/parse_cropwise.py: instead of reading CropWise's per-field
export, we GENERATE the same shape from our own field_treatments — a 'Паспорт
поля' sheet + one 'Технические операции YYYY' sheet per season, each split into
the four sections (Обработка почвы / Сев / Внесение удобрений / Внесение СЗР).
So data logged in the bot can come back out in the report format the farm knows.

Used by the /export bot command (sends the .xlsx as a Telegram document) and as a
CLI:  python -m catalog.export_multiprotocol --field 119 [out.xlsx]
"""
import argparse
import asyncio
import io
import re
import sys

import openpyxl
from openpyxl.styles import Font
from sqlalchemy import text

from bot.db import engine, resolve_field_id

# op_category → section label + the material column header (mirrors parse_cropwise)
SECTIONS = [
    ("tillage", "Обработка почвы"),
    ("sowing", "Сев"),
    ("fertilizer", "Внесение удобрений"),
    ("protection", "Внесение СЗР"),
]
HEADERS = ["№", "Дата", "Культура", "Операция", "Препарат", "Д.в.", "Норма",
           "Площадь, га", "Примечание"]


def _name_parts(name: str):
    """'Поле 119 · Хлевище' -> ('119', 'Хлевище'); 'Поле 76/108' -> ('76/108','')."""
    base = re.sub(r"^Поле\s+", "", name or "").strip()
    if " · " in base:
        num, grp = base.split(" · ", 1)
        return num.strip(), grp.strip()
    return base, ""


async def _gather(field_id):
    async with engine.connect() as conn:
        field = (await conn.execute(text(
            "SELECT name, crop, area_ha FROM fields WHERE id=:i"), {"i": field_id})).mappings().first()
        ops = (await conn.execute(text(
            "SELECT season, treatment_date, crop, operation, op_category, product, "
            "active_substance, dose, area_ha, note FROM field_treatments WHERE field_id=:i "
            "ORDER BY season, treatment_date"), {"i": field_id})).mappings().all()
        crops = (await conn.execute(text(
            "SELECT year, crop, variety FROM field_crops WHERE field_id=:i ORDER BY year DESC"),
            {"i": field_id})).all()
    return field, ops, crops


def _build_xlsx(field, ops, crops) -> bytes:
    num, grp = _name_parts(field["name"])
    wb = openpyxl.Workbook()
    bold = Font(bold=True)

    # --- Паспорт поля ---
    ws = wb.active
    ws.title = "Паспорт поля"
    ws["A1"] = "Паспорт поля"; ws["A1"].font = bold
    rows = [("Поле", num), ("Группа полей", grp),
            ("Площадь, га", float(field["area_ha"]) if field["area_ha"] is not None else ""),
            ("Текущая культура", field["crop"] or "")]
    r = 3
    for k, v in rows:
        ws.cell(r, 1, k).font = bold; ws.cell(r, 2, v); r += 1
    if crops:
        r += 1
        ws.cell(r, 1, "Севооборот:").font = bold; r += 1
        for yr, crop, variety in crops:
            ws.cell(r, 1, yr); ws.cell(r, 2, f"{crop or ''}" + (f" · {variety}" if variety else "")); r += 1

    # --- one sheet per season ---
    seasons = sorted({o["season"] for o in ops if o["season"]}, reverse=True)
    for season in seasons:
        sh = wb.create_sheet(f"Технические операции {season}")
        row = 1
        for cat, label in SECTIONS:
            cat_ops = [o for o in ops if o["season"] == season and o["op_category"] == cat]
            if not cat_ops:
                continue
            sh.cell(row, 1, label).font = bold; row += 1
            for c, h in enumerate(HEADERS, 1):
                sh.cell(row, c, h).font = bold
            row += 1
            for i, o in enumerate(cat_ops, 1):
                d = o["treatment_date"]
                vals = [i, d.strftime("%d.%m.%Y") if d else "", o["crop"] or "",
                        o["operation"] or "", o["product"] or "", o["active_substance"] or "",
                        o["dose"] or "",
                        float(o["area_ha"]) if o["area_ha"] is not None else "",
                        o["note"] or ""]
                for c, v in enumerate(vals, 1):
                    sh.cell(row, c, v)
                row += 1
            row += 1  # blank line between sections

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def build_multiprotocol(field_query, farm_id=None):
    """Resolve the field and build its multiprotocol .xlsx. Returns
    (filename, bytes) or None if the field isn't found."""
    fid = await resolve_field_id(field_query, farm_id)
    if fid is None:
        return None
    field, ops, crops = await _gather(fid)
    if field is None:
        return None
    data = _build_xlsx(field, ops, crops)
    num, _ = _name_parts(field["name"])
    safe = re.sub(r"[^0-9A-Za-z]+", "_", num).strip("_") or str(fid)
    return f"field_{safe}_multiprotocol.xlsx", data


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    ap.add_argument("--field", required=True, help="field number/name, e.g. 119 or 76/108")
    ap.add_argument("out", nargs="?", help="output .xlsx path (default: the generated name)")
    args = ap.parse_args()
    res = asyncio.run(build_multiprotocol(args.field))
    if not res:
        print(f"field not found: {args.field}", file=sys.stderr); return 1
    fname, data = res
    out = args.out or fname
    with open(out, "wb") as f:
        f.write(data)
    print(f"wrote {out} ({len(data)} bytes)", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
