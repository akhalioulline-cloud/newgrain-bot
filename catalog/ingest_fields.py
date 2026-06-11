"""Ingest the CropWise whole-farm fields export (#fields-*.xlsx) into the fields
table (+ field_crops rotation).

The 3 pilot fields are matched by leading-number + area and enriched (field_group)
rather than duplicated; the rest are inserted is_pilot=false as "Поле <name> · <group>".
field_crops gets one row per (field, year) for the years present (2025-2027):
crop, variety, sow/harvest dates, yield. Idempotent (re-match by name).

Run: docker compose -f docker-compose.prod.yml run --rm -T -v /tmp/fields.xlsx:/data.xlsx \
       bot python -m catalog.ingest_fields /data.xlsx
"""
import argparse
import asyncio
import re
import sys
from datetime import date, datetime

import openpyxl
from sqlalchemy import text

from bot.db import engine


def _num(v):
    if v in (None, ""):
        return None
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except ValueError:
        return None


def _date(v):
    if v in (None, ""):
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    for f in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(v).strip(), f).date()
        except ValueError:
            pass
    return None


def _lead_int(name):
    m = re.match(r"\D*(\d+)", name or "")
    return int(m.group(1)) if m else None


def parse(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Fields"] if "Fields" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    hdr = {str(c).strip(): i for i, c in enumerate(rows[0]) if c is not None}

    def col(r, name):
        i = hdr.get(name)
        return r[i] if (i is not None and i < len(r)) else None

    out = []
    for r in rows[1:]:
        if not any(c is not None and str(c).strip() for c in r):
            continue
        name = str(col(r, "Имя")).strip() if col(r, "Имя") is not None else ""
        if not name:
            continue
        rec = {"group": (col(r, "Группа полей") or "").strip(), "name": name,
               "area": _num(col(r, "Офиц. площ., га")) or _num(col(r, "Расч. площ., га")),
               "crops": {}}
        for y in (2025, 2026, 2027):
            crop = (col(r, f"Культура (Название) | {y}") or "").strip()
            if crop:
                rec["crops"][y] = {
                    "crop": crop,
                    "variety": (col(r, f"Сорт | {y}") or "").strip() or None,
                    "sow": _date(col(r, f"Дата сева | {y}")),
                    "harvest": _date(col(r, f"Дата уборки | {y}")),
                    "yield": _num(col(r, f"Урож., ц/га | {y}")),
                }
        out.append(rec)
    wb.close()
    return out


async def load(recs):
    ins = upd = ncrops = 0
    async with engine.begin() as conn:
        farm_id = (await conn.execute(text("SELECT farm_id FROM fields WHERE is_pilot LIMIT 1"))).scalar() \
            or (await conn.execute(text("SELECT id FROM farms ORDER BY id LIMIT 1"))).scalar()
        existing = (await conn.execute(text("SELECT id, name, area_ha FROM fields"))).all()
        by_name = {nm: fid for fid, nm, _a in existing}
        by_numarea = {}
        for fid, nm, area in existing:
            li = _lead_int(nm)
            if li is not None:
                by_numarea[(li, round(float(area)) if area is not None else None)] = fid

        for rec in recs:
            display = f"Поле {rec['name']} · {rec['group']}"
            fid = by_name.get(display)
            if not fid:  # pilot match by number+area
                fid = by_numarea.get((_lead_int(rec["name"]), round(rec["area"]) if rec["area"] else None))
            if fid:
                await conn.execute(text("UPDATE fields SET field_group=:g WHERE id=:i"),
                                   {"g": rec["group"], "i": fid})
                upd += 1
            else:
                cur = (rec["crops"].get(2026) or rec["crops"].get(2025) or {}).get("crop")
                fid = (await conn.execute(text(
                    "INSERT INTO fields (farm_id, name, crop, area_ha, is_pilot, field_group) "
                    "VALUES (:f,:n,:cr,:a,false,:g) RETURNING id"),
                    {"f": farm_id, "n": display, "cr": cur, "a": rec["area"], "g": rec["group"]})).scalar()
                ins += 1
            for y, cd in rec["crops"].items():
                await conn.execute(text(
                    "INSERT INTO field_crops (field_id, year, crop, variety, sow_date, harvest_date, yield_cwt) "
                    "VALUES (:i,:y,:cr,:v,:s,:h,:yl) ON CONFLICT (field_id, year) DO UPDATE SET "
                    "crop=EXCLUDED.crop, variety=EXCLUDED.variety, sow_date=EXCLUDED.sow_date, "
                    "harvest_date=EXCLUDED.harvest_date, yield_cwt=EXCLUDED.yield_cwt"),
                    {"i": fid, "y": y, "cr": cd["crop"], "v": cd["variety"], "s": cd["sow"],
                     "h": cd["harvest"], "yl": cd["yield"]})
                ncrops += 1
    return ins, upd, ncrops


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    args = ap.parse_args()
    recs = parse(args.xlsx)
    print(f"parsed {len(recs)} fields", file=sys.stderr)
    if not recs:
        return 1
    ins, upd, nc = asyncio.run(load(recs))
    print(f"inserted {ins} new fields, updated {upd} existing (pilots matched), "
          f"{nc} crop-year rows", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
